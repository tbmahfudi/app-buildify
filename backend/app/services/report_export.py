"""
Report export service for generating various output formats.
"""
import csv
import json
import io
from typing import List, Dict, Any
from datetime import datetime
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ReportExporter:
    """Service for exporting reports to various formats."""

    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], columns: List[str]) -> bytes:
        """Export report data to CSV format."""
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        writer.writerows(data)

        return output.getvalue().encode('utf-8')

    @staticmethod
    def export_to_json(data: List[Dict[str, Any]]) -> bytes:
        """Export report data to JSON format."""
        return json.dumps(data, indent=2, default=str).encode('utf-8')

    @staticmethod
    def export_to_excel_raw(data: List[Dict[str, Any]], columns: List[str]) -> bytes:
        """Export report data to Excel (unformatted)."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Report Data"

        # Write headers
        for col_idx, column in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=column)

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, column in enumerate(columns, start=1):
                value = row_data.get(column, '')
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_to_excel_formatted(
        data: List[Dict[str, Any]],
        columns: List[str],
        report_name: str = "Report",
        formatting_rules: List[Dict[str, Any]] = None
    ) -> bytes:
        """Export report data to Excel (formatted)."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = report_name[:31]  # Excel sheet name limit

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers with formatting
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write data with formatting
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, column in enumerate(columns, start=1):
                value = row_data.get(column, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Apply conditional formatting if rules exist
                if formatting_rules:
                    for rule in formatting_rules:
                        if column in rule.get('applies_to', []):
                            # Simple condition evaluation (expand as needed)
                            # This is a placeholder for more complex condition logic
                            pass

        # Auto-adjust column widths
        for col_idx, column in enumerate(columns, start=1):
            max_length = len(column)
            for row_idx in range(2, len(data) + 2):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_to_pdf(
        data: List[Dict[str, Any]],
        columns: List[str],
        report_name: str = "Report",
        parameters: Dict[str, Any] = None
    ) -> bytes:
        """Export report data to PDF format."""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        normal_style = styles['Normal']

        # Title
        title = Paragraph(report_name, title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Parameters (if any)
        if parameters:
            param_text = "Parameters: " + ", ".join([f"{k}: {v}" for k, v in parameters.items()])
            elements.append(Paragraph(param_text, normal_style))
            elements.append(Spacer(1, 12))

        # Generated date
        gen_date = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        elements.append(Paragraph(gen_date, normal_style))
        elements.append(Spacer(1, 20))

        # Prepare table data
        table_data = [columns]  # Headers
        for row in data:
            table_data.append([str(row.get(col, '')) for col in columns])

        # Create table
        table = Table(table_data)

        # Table styling
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_to_html(
        data: List[Dict[str, Any]],
        columns: List[str],
        report_name: str = "Report",
        parameters: Dict[str, Any] = None
    ) -> bytes:
        """Export report data to HTML format (print-friendly)."""
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{report_name}</title>
    <style>
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
        }}
        h1 {{
            color: #333;
            border-bottom: 2px solid #366092;
            padding-bottom: 10px;
        }}
        .metadata {{
            color: #666;
            font-size: 14px;
            margin-bottom: 20px;
        }}
        table {{
            border-collapse: collapse;
            width: 100%;
            margin-top: 20px;
        }}
        th {{
            background-color: #366092;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
        }}
        td {{
            border: 1px solid #ddd;
            padding: 10px;
        }}
        tr:nth-child(even) {{
            background-color: #f2f2f2;
        }}
        tr:hover {{
            background-color: #e0e0e0;
        }}
        @media print {{
            button {{
                display: none;
            }}
        }}
    </style>
</head>
<body>
    <h1>{report_name}</h1>
    <div class="metadata">
"""
        # Add parameters
        if parameters:
            html += "        <p><strong>Parameters:</strong> "
            html += ", ".join([f"{k}: {v}" for k, v in parameters.items()])
            html += "</p>\n"

        # Add generated date
        html += f"        <p><strong>Generated:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}</p>\n"
        html += f"        <p><strong>Total Records:</strong> {len(data)}</p>\n"
        html += "    </div>\n"

        # Print button
        html += "    <button onclick=\"window.print()\">Print Report</button>\n"

        # Table
        html += "    <table>\n"
        html += "        <thead>\n"
        html += "            <tr>\n"
        for col in columns:
            html += f"                <th>{col}</th>\n"
        html += "            </tr>\n"
        html += "        </thead>\n"
        html += "        <tbody>\n"

        for row in data:
            html += "            <tr>\n"
            for col in columns:
                value = row.get(col, '')
                html += f"                <td>{value}</td>\n"
            html += "            </tr>\n"

        html += "        </tbody>\n"
        html += "    </table>\n"
        html += "</body>\n"
        html += "</html>"

        return html.encode('utf-8')

    @staticmethod
    def export_report(
        data: List[Dict[str, Any]],
        columns: List[str],
        export_format: str,
        report_name: str = "Report",
        parameters: Dict[str, Any] = None,
        formatting_rules: List[Dict[str, Any]] = None
    ) -> tuple[bytes, str]:
        """
        Export report to specified format.
        Returns tuple of (file_content, file_extension)
        """
        if export_format == "csv":
            return ReportExporter.export_to_csv(data, columns), "csv"
        elif export_format == "json":
            return ReportExporter.export_to_json(data), "json"
        elif export_format == "excel_raw":
            return ReportExporter.export_to_excel_raw(data, columns), "xlsx"
        elif export_format == "excel_formatted":
            return ReportExporter.export_to_excel_formatted(
                data, columns, report_name, formatting_rules
            ), "xlsx"
        elif export_format == "pdf":
            return ReportExporter.export_to_pdf(data, columns, report_name, parameters), "pdf"
        elif export_format == "html":
            return ReportExporter.export_to_html(data, columns, report_name, parameters), "html"
        else:
            raise ValueError(f"Unsupported export format: {export_format}")
